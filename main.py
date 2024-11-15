import requests
import openpyxl
from datetime import datetime, timedelta
from sklearn.linear_model import LinearRegression
from colorama import init, Fore, Back
import os
import itertools
from statsmodels.tsa.arima.model import ARIMA

# Ensure the 'reports' folder exists for saving the Excel spreadsheets
if not os.path.exists('reports'):
    os.mkdir('reports')

# Initialize colorama for colored terminal output
init()

# Greet the user
print(Fore.GREEN + "Welcome to the Cryptocurrency Price Prediction Tool!" + Fore.RESET)
print("This tool allows you to select a cryptocurrency, analyze its historical data, and predict its future prices.")
print("Let's get started!\n")

# Define the CryptoCompare API endpoint for historical data
CRYPTOCOMPARE_API_ENDPOINT = "https://min-api.cryptocompare.com/data/v2/histohour"

# Initialize the machine learning model (Linear Regression by default)
model = LinearRegression()

# Set up an Excel workbook to store the price predictions
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Price Predictions"
ws.cell(1, 1, value="Date")
ws.cell(1, 2, value="Price")

# Set the 'again' flag to False to trigger the initial cryptocurrency selection prompt
again = False

# Start the main program loop
while True:
    # Clear previous prediction data from the sheet for a fresh start
    ws.delete_rows(2, ws.max_row)
    
    # Prompt the user to analyze another cryptocurrency if this isn't the first iteration
    if again:
        print("\nDo you want to analyze another cryptocurrency? (y/n)")
        choice = input("> ")
        if choice.lower() == "n":
            print(Fore.CYAN + "\nThank you for using the Cryptocurrency Price Prediction Tool! Goodbye!" + Fore.RESET)
            break
        elif choice.lower() != "y":
            print(Fore.RED + "Invalid input. Please enter 'y' or 'n'." + Fore.RESET)
            continue
    
    # Set the 'again' flag to True to avoid showing the initial prompt again
    else:
        again = True

    # Fetch and display the list of top cryptocurrencies by market cap
    print("\nChoose a cryptocurrency to analyze:")
    try:
        # Fetch the list from the CryptoCompare API
        response = requests.get("https://min-api.cryptocompare.com/data/top/mktcapfull?limit=60&tsym=USD")
        response.raise_for_status()
        # Extract cryptocurrency names from the response
        cryptocurrencies = [c["CoinInfo"]["Name"] for c in response.json()["Data"]]
        
        # Display cryptocurrencies with indices for selection
        for i, c in enumerate(cryptocurrencies):
            print(f"{i + 1}. {c}")

        # Handle user input for cryptocurrency selection
        while True:
            try:
                choice = int(input("> "))
                if 1 <= choice <= len(cryptocurrencies):
                    selected_cryptocurrency = cryptocurrencies[choice - 1]
                    break
                else:
                    print(Fore.RED + f"Invalid choice. Please enter a number between 1 and {len(cryptocurrencies)}." + Fore.RESET)
            except ValueError:
                print(Fore.RED + "Invalid input. Please enter a valid number." + Fore.RESET)

    except requests.RequestException:
        print(Fore.RED + "Error fetching cryptocurrency data. Please check your network connection and try again." + Fore.RESET)
        continue

    # Prompt the user to select a time period for analysis
    print("\nChoose a time period to analyze:")
    print("1. 24 hours")
    print("2. 7 days")
    print("3. 12 months")
    while True:
        try:
            time_period_choice = int(input("> "))
            if time_period_choice == 1:
                name_time_period_choice = "24 hours"
                CRYPTOCOMPARE_API_PARAMS = {"fsym": selected_cryptocurrency, "tsym": "USD", "limit": 24, "aggregate": 1}
                break
            elif time_period_choice == 2:
                name_time_period_choice = "7 days"
                CRYPTOCOMPARE_API_PARAMS = {"fsym": selected_cryptocurrency, "tsym": "USD", "limit": 168, "aggregate": 1}
                break
            elif time_period_choice == 3:
                name_time_period_choice = "12 months"
                CRYPTOCOMPARE_API_PARAMS = {"fsym": selected_cryptocurrency, "tsym": "USD", "limit": 365, "aggregate": 1}
                break
            else:
                print(Fore.RED + "Invalid choice. Please enter 1, 2, or 3." + Fore.RESET)
        except ValueError:
            print(Fore.RED + "Invalid input. Please enter a valid number." + Fore.RESET)

    # Prompt the user to choose an algorithm for price prediction
    print("\nChoose an algorithm to use for price prediction:")
    print("1. Linear Regression")
    print("2. ARIMA")
    while True:
        try:
            algorithm_choice = int(input("> "))
            if algorithm_choice in [1, 2]:
                break
            else:
                print(Fore.RED + "Invalid choice. Please enter 1 or 2." + Fore.RESET)
        except ValueError:
            print(Fore.RED + "Invalid input. Please enter a valid number." + Fore.RESET)

    # Set the selected algorithm
    if algorithm_choice == 1:
        model = LinearRegression()
    elif algorithm_choice == 2:
        # Generate potential pdq combinations for ARIMA model
        pdq = list(itertools.product(range(0, 6), range(0, 2), range(0, 2)))

    # Fetch historical price data for the selected cryptocurrency and time period
    response = requests.get(CRYPTOCOMPARE_API_ENDPOINT, params=CRYPTOCOMPARE_API_PARAMS)
    if not response.ok:
        print(Fore.RED + "Could not retrieve historical price data. Please try again." + Fore.RESET)
        continue

    history_data = response.json()["Data"]["Data"]
    if not history_data:
        print(Fore.RED + "Historical price data is empty. Please try again." + Fore.RESET)
        continue

    # Prepare the historical data for machine learning
    timestamps = [[datetime.fromtimestamp(h["time"]).timestamp() * 1000] for h in history_data]
    prices = [h["close"] for h in history_data]
    model.fit(timestamps, prices)  # Train the model on the historical price data

    # Generate future price predictions based on the chosen time period
    current_time = datetime.now()
    if time_period_choice == 1:  # Predict next 24 hours
        for i in range(24):
            next_time = current_time + timedelta(hours=i+1)
            next_timestamp = int(next_time.timestamp() * 1000)
            next_price = model.predict([[next_timestamp]])[0]
            ws.cell(i+2, 1, value=next_time)
            ws.cell(i+2, 2, value=next_price)
            ws.cell(i+2, 1).number_format = "mm/dd/yyyy hh:mm"
    elif time_period_choice == 2:  # Predict next 7 days
        for i in range(7):
            next_time = current_time + timedelta(days=i+1)
            next_timestamp = int(next_time.timestamp() * 1000)
            next_price = model.predict([[next_timestamp]])[0]
            ws.cell(i+2, 1, value=next_time)
            ws.cell(i+2, 2, value=next_price)
            ws.cell(i+2, 1).number_format = "mm/dd/yyyy hh:mm"
    elif time_period_choice == 3:  # Predict next 12 months
        for i in range(12):
            next_date_time = current_time + timedelta(days=30*(i+1))
            next_timestamp = int(next_date_time.timestamp() * 1000)
            next_price = model.predict([[next_timestamp]])[0]
            ws.cell(i+2, 1, value=next_date_time.date())
            ws.cell(i+2, 2, value=next_price)

    # Save the Excel file with a timestamped filename
    date_time_string = datetime.now().strftime("%m%d%Y-%H%M%S")
    try:
        ws.cell(1, 3, value="Algorithm")
        ws.cell(2, 3, value="Linear-Regression" if algorithm_choice == 1 else "ARIMA")
        file_path = f"reports/{'Linear-Regression' if algorithm_choice == 1 else 'ARIMA'}-{name_time_period_choice.replace(' ', '-')}-{selected_cryptocurrency}-{date_time_string}.xlsx"
        wb.save(file_path)
        print(Fore.GREEN + f"\nSaved to: {file_path}" + Fore.RESET)
    except Exception as e:
        print(Fore.RED + "Error saving to Excel:", e + Fore.RESET)
        exit()

    # Display predictions to the user in a formatted table
    print(f"\nPrice predictions for: {selected_cryptocurrency}")
    print(f"Algorithm: {'Linear Regression' if algorithm_choice == 1 else 'ARIMA'}")
    print(f"Time period: {name_time_period_choice}")
    print("+" + "-"*27 + "+" + "-"*12 + "+")
    print("| Date".ljust(27) + " | Price".ljust(14) + "|")
    print("+" + "-"*27 + "+" + "-"*12 + "+")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        date = row[0]
        price = f"${row[1]:.2f}"
        print(f"| {date.strftime('%m/%d/%Y %I:%M:%S %p').ljust(25)} | {price.rjust(10)} |")
    print("+" + "-"*27 + "+" + "-"*12 + "+")

print(Fore.CYAN + "\nThank you for using the Cryptocurrency Price Prediction Tool! Have a great day!" + Fore.RESET)
